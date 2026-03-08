"""
Grameen AI — Rural Community Voice Agent
─────────────────────────────────────────
STT  : Sarvam saaras:v3   (auto-detect language)
LLM  : AWS Haiku 4.5 (swappable — see commented AWS block)
TTS  : Sarvam bulbul:v3   (language-matched from STT)
Files: LiveKit Byte Streams on topic "document-upload"
       images → Sarvam Vision OCR
       PDFs   → pypdf text layer → Vision fallback
       Excel  → openpyxl
       CSV    → stdlib csv
RPC  : "reanalyze-document" — re-trigger spoken analysis with custom focus
Search: Exa MCP
"""

from datetime import date
import asyncio
import base64
import io
import json
import logging
import os

import httpx
from dotenv import load_dotenv
from livekit import agents, rtc
from livekit.agents import AgentServer, AgentSession, Agent, mcp
from livekit.agents.voice import room_io
from livekit.plugins import silero, sarvam, openai
from livekit.plugins import aws   # ← uncomment to switch LLM

load_dotenv(".env")
logger = logging.getLogger("grameen-ai")

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────

SARVAM_API_KEY    = os.environ["SARVAM_API_KEY"]
EXA_API_KEY       = os.environ["EXA_API_KEY"]
SARVAM_VISION_URL = "https://api.sarvam.ai/v1/vision"   # verify against latest docs

MAX_IMAGE_PX  = 1024
MAX_PDF_PAGES = 6

BYTE_STREAM_TOPIC = "document-upload"
RPC_REANALYZE     = "reanalyze-document"

# ─────────────────────────────────────────────────────────────────────────────
# Language map
# ─────────────────────────────────────────────────────────────────────────────

LANG_MAP: dict[str, str] = {
    "hi": "hi-IN", "ta": "ta-IN", "te": "te-IN", "kn": "kn-IN",
    "ml": "ml-IN", "mr": "mr-IN", "gu": "gu-IN", "bn": "bn-IN",
    "pa": "pa-IN", "or": "or-IN", "en": "en-IN", "unknown": "en-IN",
}

# ─────────────────────────────────────────────────────────────────────────────
# System prompt
# ─────────────────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = f"""You are a helpful community assistant for people in rural India. Today is {date.today()}.

TOOLS:
- You have access to a web search tool powered by Exa. Use it whenever you need current information such as:
  - Today's mandi prices or crop rates
  - Latest government scheme updates, deadlines, or new notifications
  - Current weather or pest outbreak alerts
  - Current news relevant to farmers, rural health, or government policies
  - Any fact you are not certain about
- Always search before answering questions about prices, laws, recent events, or news.
- Never say "I am searching" or mention tool calls out loud. Search silently and speak only the result.

DOCUMENT READING — when a user sends a file or document:
- You will receive the extracted text from the document automatically.
- First, in one short sentence, tell the user what kind of document it appears to be.
- Then read out the most important information clearly, in natural spoken order.
- For government forms: name the scheme or department, then read key fields like name, date, amount, status.
- For land records: state the plot number, owner name, area, and any dues or remarks.
- For Aadhaar or ID cards: confirm name, date of birth, and ID type only — never read the full ID number aloud.
- For bank documents: state account type, balance if shown, and any important transactions.
- For health documents: state the diagnosis or prescription summary — always add "please confirm with your doctor."
- For Excel or CSV files: summarise what the table contains and highlight key numbers or trends.
- If fields are blank or illegible, say so plainly — do not guess.
- Never read out long codes, reference numbers, file paths, or raw HTML verbatim — always summarise.
- If the document is in another language, say so and read what you can understand.

VOICE RULES — follow these strictly:
- Speak in short, clear sentences. Never use bullet points, asterisks, or markdown.
- Never say things like "Here are three points" or use numbered lists out loud. Weave information naturally into speech.
- Pause naturally. Use commas and full stops to create breath points.
- Keep every response under 60 words unless reading a document or the user asks for more detail.
- Never narrate your thinking process. Just speak the answer.

LANGUAGE:
- Use simple, everyday words. Avoid jargon. If you must use a technical term, explain it immediately.
- Match the user's language automatically — if they speak Hindi, respond in Hindi. If Telugu, respond in Telugu.

YOUR JOB:
- Agriculture: crop cycles, pest control, soil care, basic animal health.
- Government schemes: PM-KISAN, MGNREGA, subsidies, pensions. Exact documents and where to submit.
- Basic health: hygiene, first aid, prevention. Always refer to PHC, ASHA worker, or local doctor for illness.
- Money and education: safe banking, microloans, mobile payments, children's learning resources.
- Current news: agriculture news, rural policy changes, weather alerts, market updates relevant to rural India.
- Documents: read and explain any file the user sends — forms, IDs, land records, prescriptions, bank statements.

HARD LIMITS:
- Never diagnose illness or suggest medicine.
- Never advise on risky investments.
- If you do not know a real-time price, law, or weather detail, search first, then answer honestly if nothing is found.
- For step-by-step processes, speak each step as a natural sentence in sequence.
- Never reveal the full Aadhaar number, bank account number, or any sensitive ID verbatim.

"""
# ─────────────────────────────────────────────────────────────────────────────
# File processor
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# Dependencies: pip install sarvamai beautifulsoup4
# ─────────────────────────────────────────────────────────────────────────────

class FileProcessor:
    """
    Routes raw bytes → plain text for the LLM.

    PDF / image/*          → Sarvam Document Intelligence (job-based OCR)
    application/vnd.*excel → openpyxl
    text/csv               → stdlib csv
    anything else          → Sarvam Document Intelligence best-effort
    """

    def __init__(self):
        from sarvamai import SarvamAI
        self._client = SarvamAI(api_subscription_key=SARVAM_API_KEY)

    async def process(self, data: bytes, mime_type: str, filename: str) -> str:
        mime = mime_type.lower()
        try:
            if mime.startswith("image/") or mime == "application/pdf":
                return await self._process_via_doc_intelligence(data, filename)
            elif "spreadsheet" in mime or "excel" in mime or mime == "application/vnd.ms-excel":
                return await asyncio.to_thread(self._process_excel, data)
            elif mime in ("text/csv", "application/csv"):
                return await asyncio.to_thread(self._process_csv, data)
            else:
                logger.warning("Unknown mime %s for '%s' — trying Doc Intelligence", mime, filename)
                return await self._process_via_doc_intelligence(data, filename)
        except Exception as exc:
            logger.exception("FileProcessor error for '%s'", filename)
            return f"[Error processing '{filename}': {exc}]"

    # ── Sarvam Document Intelligence ──────────────────────────────────────

    async def _process_via_doc_intelligence(self, data: bytes, filename: str) -> str:
        return await asyncio.to_thread(self._doc_intelligence_sync, data, filename)

    def _doc_intelligence_sync(self, data: bytes, filename: str) -> str:
        import tempfile, zipfile, os
        from bs4 import BeautifulSoup

        # Write bytes to a temp file — SDK needs a file path
        suffix = os.path.splitext(filename)[-1] or ".pdf"
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp_in:
            tmp_in.write(data)
            tmp_path = tmp_in.name

        tmp_zip = tmp_path + "_output.zip"
        tmp_out_dir = tmp_path + "_out"

        try:
            # Create job
            job = self._client.document_intelligence.create_job(
                language="en-IN",
                output_format="html",
            )
            logger.info("Sarvam Doc Intelligence job created: %s for '%s'", job.job_id, filename)

            # Upload + start
            job.upload_file(tmp_path)
            job.start()

            # Wait (blocks thread — wrapped in asyncio.to_thread above)
            status = job.wait_until_complete()
            logger.info("Doc Intelligence job %s completed: %s", job.job_id, status.job_state)

            if str(status.job_state).lower() not in ("completed", "success", "done"):
                return f"[Document processing failed: job state = {status.job_state}]"

            # Download ZIP output
            job.download_output(tmp_zip)

            # Extract HTML from ZIP and parse to plain text
            os.makedirs(tmp_out_dir, exist_ok=True)
            with zipfile.ZipFile(tmp_zip, "r") as zf:
                zf.extractall(tmp_out_dir)

            # Collect all HTML files and extract text
            parts = []
            for root, _, files in os.walk(tmp_out_dir):
                for f in sorted(files):
                    if f.endswith(".html") or f.endswith(".htm"):
                        with open(os.path.join(root, f), "r", encoding="utf-8", errors="replace") as fh:
                            soup = BeautifulSoup(fh.read(), "html.parser")
                            text = soup.get_text(separator="\n", strip=True)
                            if text:
                                parts.append(text)

            return "\n\n".join(parts) if parts else "[Document processed but no text extracted]"

        finally:
            # Clean up temp files
            for path in [tmp_path, tmp_zip]:
                try:
                    os.unlink(path)
                except Exception:
                    pass
            try:
                import shutil
                shutil.rmtree(tmp_out_dir, ignore_errors=True)
            except Exception:
                pass

    # ── Excel ──────────────────────────────────────────────────────────────

    @staticmethod
    def _process_excel(data: bytes) -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            parts = []
            for name in wb.sheetnames:
                rows = [
                    " | ".join(str(c) if c is not None else "" for c in row)
                    for row in wb[name].iter_rows(values_only=True)
                    if any(c is not None for c in row)
                ]
                if rows:
                    parts.append(f"Sheet: {name}\n" + "\n".join(rows))
            return "\n\n".join(parts) or "[Excel is empty]"
        except ImportError:
            return "[openpyxl not installed]"

    # ── CSV ────────────────────────────────────────────────────────────────

    @staticmethod
    def _process_csv(data: bytes) -> str:
        import csv
        rows = [
            " | ".join(row)
            for row in csv.reader(io.StringIO(data.decode("utf-8", errors="replace")))
            if any(row)
        ]
        return "\n".join(rows) or "[CSV is empty]"

# ─────────────────────────────────────────────────────────────────────────────
# Byte stream handler
# ─────────────────────────────────────────────────────────────────────────────
# Follows the LiveKit Python pattern exactly:
#   - async_handle reads the stream and does work
#   - sync handle_byte_stream wraps it in a task (required by the SDK)
#   - tasks stored in a set to prevent GC before completion
# ─────────────────────────────────────────────────────────────────────────────

_processor = FileProcessor()
_active_tasks: set[asyncio.Task] = set()


def register_byte_stream_handler(room: rtc.Room, session: AgentSession) -> None:

    async def _handle(reader, participant_identity: str) -> None:
        info      = reader.info
        # reader.info properties: name, mime_type, topic, timestamp, id, size
        filename  = info.name      if hasattr(info, "name")      else info.get("name", "document")
        mime_type = info.mime_type if hasattr(info, "mime_type") else info.get("mime_type", "application/octet-stream")

        logger.info("Receiving '%s' (%s) from %s", filename, mime_type, participant_identity)

        # Acknowledge receipt immediately — don't leave the user in silence
        # await session.generate_reply(
        #         user_input=f"Tell the user in one short sentence that you received their document and are reading it. document filename: {filename}"
        # )

        # Collect all chunks (mirrors the LiveKit Python example pattern)
        chunks: list[bytes] = []
        async for chunk in reader:
            chunks.append(chunk)
        file_bytes = b"".join(chunks)

        logger.info("Received %d bytes for '%s'", len(file_bytes), filename)

        extracted = await _processor.process(file_bytes, mime_type, filename)

        # ── Persist document in chat history ─────────────────────────────
        # chat_ctx.append() adds this to the LLM's conversation history
        # permanently, so follow-up questions ("what was the name on that
        # form?") work in all future turns — not just the immediate reply.
        chat_ctx = session.current_agent.chat_ctx.copy()
        chat_ctx.add_message(role="user", content=f"[Document uploaded: {filename}]\n\n{extracted}")
        await session.current_agent.update_chat_ctx(chat_ctx)
        # session.chat_ctx.append(
        #     role="user",
        #     text=(
        #         f"[I uploaded a document: '{filename}']\n\n"
        #         f"{extracted}"
        #     ),
        # )

        # generate_reply sees the full context including the doc above.
        await session.generate_reply(
                instructions="""
                Acknowledge that you got the document/s!
                """
        )

    def _sync_handle(reader, participant_identity: str) -> None:
        task = asyncio.create_task(_handle(reader, participant_identity))
        _active_tasks.add(task)
        task.add_done_callback(_active_tasks.discard)

    room.register_byte_stream_handler(BYTE_STREAM_TOPIC, _sync_handle)
    logger.info("Byte stream handler ready on topic '%s'", BYTE_STREAM_TOPIC)


# ─────────────────────────────────────────────────────────────────────────────
# RPC handler
# ─────────────────────────────────────────────────────────────────────────────
# Frontend calls:
#   room.localParticipant.performRpc({
#     destinationIdentity: '<agent-identity>',
#     method: 'reanalyze-document',
#     payload: JSON.stringify({ instruction: 'Focus only on expiry dates.' })
#   })

def register_rpc_methods(room: rtc.Room, session: AgentSession) -> None:

    @room.local_participant.register_rpc_method(RPC_REANALYZE)
    async def handle_reanalyze(data) -> str:
        try:
            body        = json.loads(data.payload or "{}")
            instruction = body.get("instruction", "Re-summarise the last document sent.")
        except json.JSONDecodeError:
            instruction = data.payload or "Re-summarise the last document sent."

        logger.info("RPC reanalyze from %s: %s", data.caller_identity, instruction)
        await session.generate_reply(instructions=instruction)
        return "ok"


# ─────────────────────────────────────────────────────────────────────────────
# Agent
# ─────────────────────────────────────────────────────────────────────────────

class Assistant(Agent):
    def __init__(self) -> None:
        super().__init__(instructions=SYSTEM_PROMPT)


# ─────────────────────────────────────────────────────────────────────────────
# Session handler
# ─────────────────────────────────────────────────────────────────────────────

server = AgentServer()


@server.rtc_session()
async def session_handler(ctx: agents.JobContext):

    # ── STT ─────────────────────────────────────────────────────────────────
    stt = sarvam.STT(
        language="unknown",
        model="saaras:v3",
        mode="transcribe",
        api_key=SARVAM_API_KEY,
    )

    # ── Language-aware TTS ───────────────────────────────────────────────────
    detected_lang: list[str] = ["unknown"]

    def on_transcript(transcript) -> None:
        detected_lang[0] = getattr(transcript, "language", "unknown") or "unknown"

    stt.on("transcript", on_transcript)

    def make_tts() -> sarvam.TTS:
        return sarvam.TTS(
            target_language_code=LANG_MAP.get(detected_lang[0], "en-IN"),
            model="bulbul:v3",
            speaker="shubh",
            api_key=SARVAM_API_KEY,
            pace=1.15,
        )

    # ── LLM ─────────────────────────────────────────────────────────────────
    # llm = openai.LLM(
    #     model="zai-org/glm-4.7-original:thinking",
    #     api_key=os.environ["NANO_API_KEY"],
    #     base_url=os.environ["NANO_BASE_URL"],
    #     temperature=0.4,
    # )
    # Swap to AWS Bedrock anytime:
    llm = aws.LLM(
        model="us.anthropic.claude-haiku-4-5-20251001-v1:0",
        # model="openai.gpt-oss-safeguard-120b",
        temperature=0.4,
    )

    # ── Web search ───────────────────────────────────────────────────────────
    exa_mcp = mcp.MCPServerStdio(
        command="bunx",
        args=["-y", "exa-mcp-server"],
        env={"EXA_API_KEY": EXA_API_KEY},
        client_session_timeout_seconds=120,
    )

    # ── Session ──────────────────────────────────────────────────────────────
    session = AgentSession(
        llm=llm,
        stt=stt,
        tts=make_tts(),
        vad=silero.VAD.load(),
        mcp_servers=[exa_mcp],
        # use_tts_aligned_transcript=True
    )

    # Byte stream handler can be registered before connect — it just
    # attaches a listener and doesn't touch local_participant.
    register_byte_stream_handler(ctx.room, session)

    await session.start(
        room=ctx.room,
        agent=Assistant(),
        room_input_options=room_io.RoomInputOptions(close_on_disconnect=True),
    )

    # RPC must be registered AFTER session.start() because it needs
    # room.local_participant, which only exists once the room is connected.
    register_rpc_methods(ctx.room, session)

    if ctx.room.remote_participants:
        await session.generate_reply(
            instructions="Greet the user warmly in one short sentence. Do not mention capabilities yet."
        )


if __name__ == "__main__":
    agents.cli.run_app(server)
